#!/usr/bin/env python3
"""
IEBC - Script de Actualización Automática de Tableros
=====================================================
Lee Master UOPs.xlsx y actualiza los datos embebidos en los 3 tableros HTML.
Diseñado para ejecutarse diariamente a las 2AM via Power Automate Desktop o Task Scheduler.

Uso: python actualizar_tableros.py [--ruta-excel RUTA] [--ruta-html RUTA]
Por defecto busca los archivos en el mismo directorio del script.
"""

import os
import sys
import re
import json
import struct
import logging
from datetime import datetime, date

# Configuración
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, 'Master UOPs.xlsx')
HTML_FILES = ['tablero_cfo.html', 'tablero_ceo.html', 'tablero_produccion.html']

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(SCRIPT_DIR, 'actualizacion.log'), encoding='utf-8')
    ]
)
log = logging.getLogger(__name__)


def serialize(val):
    """Serializa valores para JSON."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    return val


def repair_xlsx(path):
    """Repara archivos Excel truncados por OneDrive (sin EOCD record)."""
    with open(path, 'rb') as f:
        data = f.read()
    
    if data[-22:][0:4] == b'PK\x05\x06':
        return path  # Already OK
    
    last_cd = data.rfind(b'PK\x01\x02')
    if last_cd < 0:
        raise Exception("Archivo Excel corrupto: no se encontró directorio central ZIP")
    
    first_cd = data.find(b'PK\x01\x02')
    pos = first_cd
    count = 0
    while pos < len(data) and data[pos:pos+4] == b'PK\x01\x02':
        fname_len = struct.unpack('<H', data[pos+28:pos+30])[0]
        extra_len = struct.unpack('<H', data[pos+30:pos+32])[0]
        comment_len = struct.unpack('<H', data[pos+32:pos+34])[0]
        pos += 46 + fname_len + extra_len + comment_len
        count += 1
    
    cd_end = pos
    cd_size = cd_end - first_cd
    eocd = struct.pack('<4sHHHHIIH', b'PK\x05\x06', 0, 0, count, count, cd_size, first_cd, 0)
    
    repaired = data[:cd_end] + eocd
    rep_path = path + '.repaired.xlsx'
    with open(rep_path, 'wb') as f:
        f.write(repaired)
    
    log.info(f"Excel reparado ({count} entradas ZIP)")
    return rep_path


# ==========================================
# Mapeo de campos Excel → JavaScript
# ==========================================
FIELD_MAP_MASTER = {
    'UOP': 'uop',
    'Comitente': 'comitente',
    'Descripción UOP': 'descripcion_uop',
    'Sub Rubro': 'sub_rubro',
    'Cert Nro': 'cert_nro',
    'Concepto': 'concepto',
    'Concepto Agrupado': 'concepto_agrupado',
    'Período': 'periodo',
    'Anticipo Desacopiado ($)': 'anticipo_desacopiado',
    'Importe Certificado': 'importe_neto_cert',
    'Estado Validación por insp.': 'estado_validacion',
    'Tipo': 'tipo_comprobante',
    'N° Comprobante': 'nro_comprobante',
    'Fecha Factura': 'fecha_factura',
    'Concepto / Descripción': 'concepto_descripcion',
    'Anulada': 'anulada',
    'NC': 'nc',
    'Importe Neto Factura': 'importe_neto_factura',
    'IVA 10,05': 'iva_105',
    'IVA 21%': 'iva_21',
    'Perp IVA': 'percep_iva',
    'IIBB Bs As ($)': 'iibb_bsas',
    'IIBB CABA ($)': 'iibb_caba',
    'IIBB Tucuman($)': 'iibb_tucuman',
    'IIBB Sta Fe ($)': 'iibb_stafe',
    'Importe de factura ($)': 'importe_factura',
    'Fecha Pago 1': 'fecha_pago_1',
    'Monto Pago 1 ($)': 'monto_pago_1',
    'Fecha Pago 2': 'fecha_pago_2',
    'Monto Pago 2 ($)': 'monto_pago_2',
    'Fecha Pago 3+': 'fecha_pago_3',
    'Monto Pago 3+ ($)': 'monto_pago_3',
    'Ret ganancias': 'ret_ganancias',
    'Total Cobrado ($)': 'total_cobrado',
    'Saldo ($)': 'saldo',
    'Demora Pago (días)': 'demora_pago',
}

NUMERIC_FIELDS = {
    'importe_neto_cert', 'importe_neto_factura', 'importe_factura',
    'iva_105', 'iva_21', 'percep_iva', 'iibb_bsas', 'iibb_caba',
    'iibb_tucuman', 'iibb_stafe', 'monto_pago_1', 'monto_pago_2',
    'monto_pago_3', 'total_cobrado', 'saldo', 'demora_pago',
    'anticipo_desacopiado', 'ret_ganancias'
}

FIELD_MAP_ANTICIPOS = {
    'UOP': 'uop',
    'Descripción UOP': 'descripcion_uop',
    'Tipo Anticipo': 'tipo_anticipo',
    'Nro Anticipo': 'nro_anticipo',
    'Factura': 'factura',
    'Fecha Desembolso': 'fecha_desembolso',
    'Monto Original ($)': 'monto_original',
    'Con actualizacion': 'con_actualizacion',
    'Total Desacopiado ($)': 'total_desacopiado',
    'Saldo Pendiente ($)': 'saldo_pendiente',
    '% Desacopiado': 'pct_desacopiado',
}


def read_excel(excel_path):
    """Lee las 3 hojas del Excel y devuelve records, anticipos, mora_data."""
    import openpyxl
    
    log.info(f"Leyendo Excel: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception:
        log.warning("Excel con posible truncamiento, intentando reparar...")
        repaired = repair_xlsx(excel_path)
        wb = openpyxl.load_workbook(repaired, data_only=True)
    
    log.info(f"Hojas encontradas: {wb.sheetnames}")
    
    # ====== HOJA 1: Master ======
    ws = wb['Master']
    headers = [c.value for c in ws[1]]
    
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = {}
        for i, h in enumerate(headers):
            if h is None or h not in FIELD_MAP_MASTER:
                continue
            js_key = FIELD_MAP_MASTER[h]
            val = serialize(row[i]) if i < len(row) else None
            if val is None and js_key in NUMERIC_FIELDS:
                val = 0
            rec[js_key] = val
        records.append(rec)
    
    log.info(f"Master: {len(records)} registros")
    
    # ====== HOJA 2: Anticipos ======
    anticipos = []
    for sn in wb.sheetnames:
        if 'anticipo' in sn.lower():
            ws_ant = wb[sn]
            ant_headers = [c.value for c in ws_ant[1]]
            for row in ws_ant.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rec = {}
                for i, h in enumerate(ant_headers):
                    if h is None or h not in FIELD_MAP_ANTICIPOS:
                        continue
                    js_key = FIELD_MAP_ANTICIPOS[h]
                    val = serialize(row[i]) if i < len(row) else None
                    if val is None:
                        val = 0
                    rec[js_key] = val
                anticipos.append(rec)
            break
    
    log.info(f"Anticipos: {len(anticipos)} registros")
    
    # ====== HOJA 3: Mora (530) ======
    mora_data = None
    for sn in wb.sheetnames:
        if '530' in sn or 'mora' in sn.lower():
            ws_mora = wb[sn]
            
            # Parámetros
            tna_row = list(ws_mora.iter_rows(min_row=5, max_row=5, values_only=True))[0]
            plazo_row = list(ws_mora.iter_rows(min_row=6, max_row=6, values_only=True))[0]
            corte_row = list(ws_mora.iter_rows(min_row=7, max_row=7, values_only=True))[0]
            
            tna = float(tna_row[1]) if tna_row[1] else 0.255
            plazo = int(plazo_row[1]) if plazo_row[1] else 60
            fecha_corte = serialize(corte_row[1])
            
            # Facturas (desde fila 10)
            facturas = []
            total_cert = 0
            total_interes = 0
            total_reclamar = 0
            dias_list = []
            
            for row in ws_mora.iter_rows(min_row=10, values_only=True):
                vals = [serialize(v) for v in row]
                nro = vals[0]
                if nro is None or isinstance(nro, str):
                    continue
                
                monto_cert = float(vals[7]) if vals[7] else 0
                dias_mora = int(vals[6]) if vals[6] else 0
                tasa = float(vals[8]) if vals[8] else 0
                interes = float(vals[9]) if vals[9] else 0
                monto_total = float(vals[10]) if vals[10] else 0
                
                facturas.append({
                    "nro_factura": nro,
                    "fecha_certificado": vals[1],
                    "fecha_emision": vals[2],
                    "plazo_contractual": int(vals[3]) if vals[3] else plazo,
                    "fecha_vencimiento": vals[4],
                    "fecha_pago_efectivo": vals[5],
                    "dias_mora": dias_mora,
                    "monto_certificado": monto_cert,
                    "tasa_interes": tasa,
                    "interes_mora": interes,
                    "monto_total": monto_total
                })
                
                total_cert += monto_cert
                total_interes += interes
                total_reclamar += monto_total
                if dias_mora > 0:
                    dias_list.append(dias_mora)
            
            mora_data = {
                "parametros": {
                    "tasa_anual": tna,
                    "plazo_contractual": plazo,
                    "fecha_corte": fecha_corte,
                    "tasa_mas_3": tna + 0.03
                },
                "facturas": facturas,
                "resumen": {
                    "total_certificados": total_cert,
                    "total_interes_mora": total_interes,
                    "total_reclamar": total_reclamar,
                    "certificados_en_mora": len([f for f in facturas if f["dias_mora"] > 0 and f["monto_certificado"] > 0]),
                    "demora_promedio": round(sum(dias_list) / len(dias_list), 1) if dias_list else 0
                }
            }
            log.info(f"Mora: {len(facturas)} facturas, Total a reclamar: ${total_reclamar:,.2f}")
            break
    
    wb.close()
    return records, anticipos, mora_data


def replace_json_array(html, key, new_data):
    """Reemplaza un array JSON embebido en el HTML por key."""
    pattern = f'"{key}"\\s*:\\s*\\['
    m = re.search(pattern, html)
    if not m:
        log.warning(f"  Key '{key}' no encontrada en HTML")
        return html
    
    pos = m.end() - 1  # posición del [
    bracket_count = 0
    for i in range(pos, len(html)):
        if html[i] == '[':
            bracket_count += 1
        elif html[i] == ']':
            bracket_count -= 1
            if bracket_count == 0:
                end = i + 1
                break
    
    old_len = end - pos
    new_json = json.dumps(new_data, ensure_ascii=False, default=str)
    html = html[:pos] + new_json + html[end:]
    log.info(f"  '{key}': {old_len} → {len(new_json)} chars")
    return html


def replace_json_object(html, var_name, new_data):
    """Reemplaza un objeto JSON asignado a una variable const.
    Expects the variable to be on its own line(s), ending with ;\n before </script>."""
    pattern = f'(?:const|var)\\s+{var_name}\\s*=\\s*'
    m = re.search(pattern, html)
    if not m:
        log.warning(f"  Variable '{var_name}' no encontrada en HTML")
        return html

    start = m.end()
    # Find the next </script> tag - the JSON must end before it
    script_end = html.find('\n</script>', start)
    if script_end < 0:
        script_end = html.find('</script>', start)
    if script_end < 0:
        log.warning(f"  No se pudo encontrar </script> despues de '{var_name}'")
        return html

    # Everything from start to script_end is the JSON + semicolon
    old_chunk = html[start:script_end].rstrip().rstrip(';').rstrip()
    old_len = len(old_chunk)
    new_json = json.dumps(new_data, ensure_ascii=False)
    html = html[:start] + new_json + ';\n' + html[script_end:]
    log.info(f"  '{var_name}': {old_len} → {len(new_json)} chars")
    return html


def update_html_file(filepath, records, anticipos, mora_data=None):
    """Actualiza un archivo HTML con los datos frescos."""
    log.info(f"Actualizando: {filepath}")
    
    with open(filepath, 'r', encoding='utf-8') as f:
        html = f.read()
    
    # Reemplazar records
    html = replace_json_array(html, 'records', records)
    
    # Reemplazar anticipos
    html = replace_json_array(html, 'anticipos', anticipos)
    
    # Reemplazar MORA_DATA (solo CFO)
    if mora_data and 'MORA_DATA' in html:
        html = replace_json_object(html, 'MORA_DATA', mora_data)
    
    # Limpiar bytes nulos (por si OneDrive truncó)
    html = html.rstrip('\x00').rstrip()
    if not html.rstrip().endswith('</html>'):
        # Verificar y arreglar cierre
        if '</script>' not in html[-100:] and '<script>' in html:
            html += '\n</script>'
        if '</body>' not in html[-50:]:
            html += '\n</body>'
        html += '\n</html>'
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    
    log.info(f"  Guardado: {len(html)//1024}KB")


def push_to_github(html_dir):
    """Sube los HTML actualizados al repositorio GitHub Pages."""
    import subprocess

    REPO_URL = "https://github.com/MperezIEBC/Tableros_IEBC.git"
    REPO_DIR = os.path.join(os.path.expanduser("~"), "Tableros_IEBC_repo")

    try:
        # Clonar o actualizar repo local
        if not os.path.exists(REPO_DIR):
            log.info(f"Clonando repositorio en {REPO_DIR}...")
            subprocess.run(["git", "clone", REPO_URL, REPO_DIR], check=True, capture_output=True, text=True)
        else:
            log.info("Actualizando repo local...")
            subprocess.run(["git", "-C", REPO_DIR, "pull", "--rebase"], check=True, capture_output=True, text=True)

        # Copiar HTMLs actualizados al repo
        import shutil
        archivos_copiados = 0
        for fname in HTML_FILES + ['landing_iebc.html']:
            src = os.path.join(html_dir, fname)
            dst = os.path.join(REPO_DIR, fname)
            if os.path.exists(src):
                shutil.copy2(src, dst)
                archivos_copiados += 1

        log.info(f"  {archivos_copiados} archivos copiados al repo")

        # Git add, commit, push
        subprocess.run(["git", "-C", REPO_DIR, "add", "-A"], check=True, capture_output=True, text=True)

        # Verificar si hay cambios
        result = subprocess.run(["git", "-C", REPO_DIR, "status", "--porcelain"], capture_output=True, text=True)
        if not result.stdout.strip():
            log.info("  Sin cambios - no se necesita push")
            return

        fecha = datetime.now().strftime('%Y-%m-%d %H:%M')
        subprocess.run(["git", "-C", REPO_DIR, "commit", "-m", f"Actualización automática {fecha}"], check=True, capture_output=True, text=True)
        subprocess.run(["git", "-C", REPO_DIR, "push"], check=True, capture_output=True, text=True)

        log.info("  ✓ Push a GitHub Pages exitoso")
        log.info("  URL: https://mpereziebc.github.io/Tableros_IEBC/landing_iebc.html")

    except FileNotFoundError:
        log.warning("  Git no está instalado - no se puede hacer push a GitHub")
        log.warning("  Instalá Git desde https://git-scm.com/download/win")
    except subprocess.CalledProcessError as e:
        log.error(f"  Error en Git: {e.stderr if e.stderr else e}")
    except Exception as e:
        log.error(f"  Error push GitHub: {e}")


def main():
    """Función principal."""
    import argparse
    parser = argparse.ArgumentParser(description='Actualizar tableros IEBC desde Excel')
    parser.add_argument('--ruta-excel', default=DEFAULT_EXCEL, help='Ruta al archivo Master UOPs.xlsx')
    parser.add_argument('--ruta-html', default=SCRIPT_DIR, help='Directorio con los archivos HTML')
    parser.add_argument('--no-push', action='store_true', help='Skip GitHub push (para CI/GitHub Actions)')
    args = parser.parse_args()
    
    log.info("=" * 60)
    log.info("IEBC - Actualización automática de tableros")
    log.info(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)
    
    # Verificar Excel
    if not os.path.exists(args.ruta_excel):
        log.error(f"Excel no encontrado: {args.ruta_excel}")
        sys.exit(1)
    
    # Leer datos
    try:
        records, anticipos, mora_data = read_excel(args.ruta_excel)
    except Exception as e:
        log.error(f"Error leyendo Excel: {e}")
        sys.exit(1)
    
    # Actualizar cada HTML
    errores = 0
    for fname in HTML_FILES:
        fpath = os.path.join(args.ruta_html, fname)
        if not os.path.exists(fpath):
            log.warning(f"Archivo no encontrado: {fpath}")
            continue
        try:
            update_html_file(fpath, records, anticipos, mora_data)
        except Exception as e:
            log.error(f"Error actualizando {fname}: {e}")
            errores += 1
    
    if errores == 0:
        log.info("\n✓ Todos los tableros actualizados exitosamente")
        # Push a GitHub Pages
        if not args.no_push:
            push_to_github(args.ruta_html)
        else:
            log.info("  (Push a GitHub omitido - modo CI)")
    else:
        log.warning(f"\n⚠ {errores} error(es) durante la actualización")

    log.info("=" * 60)
    return errores


if __name__ == '__main__':
    sys.exit(main())
